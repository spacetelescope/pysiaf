"""Functions to write Science Instrument Aperture Files (SIAF).

SIAF content in an aperture_collection object can be written to an xml file that can be ingested in the PRD.
Format and order of the xml fields are defined in SIAF reference files.

Writing to Microsoft Excel .xlsx format is supported.

Writing to .csv and other formats supported by astropy.table.Table.write is enabled.


Authors
-------

    Johannes Sahlmann


"""

import numpy as np
import os

# import git
import lxml.etree as ET
from astropy.time import Time
from astropy.table import Table, Column
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Alignment

from ..version import __version__
from ..constants import _JWST_TEMPORARY_ROOT
# from ..constants import PACKAGE_VERSION, _JWST_TEMPORARY_ROOT
from ..aperture import PRD_REQUIRED_ATTRIBUTES_ORDERED, SIAF_XML_FIELD_FORMAT, FLOAT_ATTRIBUTES

# dictionary used to set field precision in SIAF.XML
xml_decimal_precision = {}
field_names = list(SIAF_XML_FIELD_FORMAT['field_name'])
for attr in PRD_REQUIRED_ATTRIBUTES_ORDERED:
    index = field_names.index(attr)
    xml_decimal_precision[attr] = SIAF_XML_FIELD_FORMAT['pyformat'][index]


def write_jwst_siaf(aperture_collection, filename=None, basepath=None, label=None, file_format='xml', verbose=True):
    """Write the content of aperture_collection into xml and xlsx files that can be delivered to
    the PRD.

    Parameters
    ----------
    aperture_collection : ApertureCollection
        dictionary of apertures
    filename
    basepath
    label
    file_format : str list
        one of ['xml', 'xlsx', 'csv', and formats supported by astropy Table.write]
    verbose

    Returns
    -------

    TODO
    ----
        test support of astropy Table.write formats (FITS not working)


    """
    if type(file_format) == str:
        file_format = [file_format]


    aperture_names = np.array([key for key in aperture_collection.apertures.keys()])
    instrument = aperture_collection.apertures[aperture_names[0]].InstrName

    if instrument == 'NIRCAM':
        instrument = 'NIRCam'
    elif instrument == 'NIRSPEC':
        instrument = 'NIRSpec'

    if (filename is not None) and (len(list(file_format)) != 1):
        raise RuntimeError('When filename is specified, only one output format is supported')

    if label is not None:
        name_seed = instrument + '_SIAF_{}'.format(label)
    else:
        name_seed = instrument + '_SIAF'

    filenames = []
    # hostname = os.uname()[1]
    username = os.getlogin()
    timestamp = Time.now()

    for file_format in list(file_format):
        if filename is None:
            if basepath is None:
                basepath = _JWST_TEMPORARY_ROOT
            if not os.path.isdir(basepath):
                raise RuntimeError("Could not write SIAF data "
                                   "to {}. Directory does not exist.".format(basepath))
            if file_format == 'xml':
                out_filename = os.path.join(basepath, name_seed+'.xml')
            elif file_format == 'xlsx':
                out_filename = os.path.join(basepath, name_seed+'.xlsx')
            # elif file_format == 'csv':
            #     out_filename = os.path.join(basepath, name_seed+'.csv')
            else:
                out_filename = os.path.join(basepath, name_seed+'.{}'.format(file_format))
        else:
            out_filename = filename

        if file_format == 'xml':
            root = ET.Element('SiafEntries')

            # add generation info as comment to SIAFXML
            root.append(ET.Comment('Generated {} {}'.format(timestamp.isot, timestamp.scale)))
            root.append(ET.Comment('by {}'.format(username)))
            # try:
            #     repo = git.Repo(os.path.abspath(__file__), search_parent_directories=True)
            #     git_version = git.Git(repo.working_dir).describe()
            #     root.append(ET.Comment('pysiaf git-version {}'.format(git_version)))
            # except git.exc.InvalidGitRepositoryError:
            root.append(ET.Comment('pysiaf version {}'.format(__version__)))

            for aperture_name in aperture_names:

                aperture = aperture_collection.apertures[aperture_name]
                siaf_entry = ET.SubElement(root, 'SiafEntry')
                for attribute in PRD_REQUIRED_ATTRIBUTES_ORDERED:
                    attribute_value = getattr(aperture_collection.apertures[aperture_name], attribute)
                    if attribute_value is None:
                        attribute_text = None

                    # NIRSpec special case
                    elif (aperture.AperType in ['TRANSFORM']) and (attribute in 'XSciRef YSciRef XSciScale YSciScale V2Ref V3Ref'.split()):
                        attribute_text = '{:{prec}}'.format(attribute_value,
                                                            prec=12).strip()
                    elif attribute in FLOAT_ATTRIBUTES:
                        attribute_text = '{:{prec}}'.format(attribute_value, prec=xml_decimal_precision[attribute]).strip()
                    else:
                        attribute_text = str(attribute_value)

                    if (not isinstance(attribute_value, str)) and (attribute_text is not None):
                        if np.isnan(attribute_value):
                            attribute_text = None

                    ET.SubElement(siaf_entry, attribute).text = attribute_text

            doc = ET.ElementTree(root)

            doc.write(out_filename, pretty_print=True, xml_declaration=False)
            if verbose:
                print('Wrote Siaf to xml file {}'.format(out_filename))

        elif file_format == 'xlsx':
            siaf_workbook = Workbook()

            ws1 = siaf_workbook.active
            ws1.title = 'SIAF'

            header_row_description = 1
            header_row_attributes = 2

            # write descriptive header
            for j, attribute_name in enumerate(PRD_REQUIRED_ATTRIBUTES_ORDERED):
                col = j + 1
                if attribute_name == 'InstrName':
                    text = 'Aperture Basic Info'
                elif attribute_name == 'XDetSize':
                    text = 'Detector Frame'
                elif attribute_name == 'XSciSize':
                    text = 'Science Frame'
                elif attribute_name == 'V2Ref':
                    text = 'V Frame'
                elif attribute_name == 'V2IdlYAngle':
                    text = 'Frame Relationships'
                elif attribute_name == 'XIdlVert1':
                    text = 'Vertices'
                elif attribute_name == 'Sci2IdlDeg':
                    text = 'Science to Ideal Polynomial'
                else:
                    text = ''

                cell = ws1.cell(column=col, row=header_row_description, value="{}".format(text))
                cell.font = Font(name='Courier', b=True, i=True, family=3.0, sz=14.0)
                # cell.font.color = Color(rgb='FF0000FF', type='rgb')

            # write aperture attributes
            for j, attribute_name in enumerate(PRD_REQUIRED_ATTRIBUTES_ORDERED):
                col = j + 1
                cell = ws1.cell(column=col, row=header_row_attributes, value="{}".format(attribute_name))
                cell.font = Font(name='Calibri', b=True, family=2.0, sz=15.0)
                cell.alignment = Alignment(horizontal='center')

            # write aperture values
            for i, aper_name in enumerate(aperture_names):
                aperture = aperture_collection.apertures[aper_name]
                # aperture = siaf[aper_name]

                row = i + 1 + header_row_attributes
                for j, attribute_name in enumerate(PRD_REQUIRED_ATTRIBUTES_ORDERED):
                    col = j + 1
                    cell = ws1.cell(column=col, row=row, value="{}".format(getattr(aperture, attribute_name)))
                    if attribute_name not in 'InstrName	AperName DDCName AperType AperShape'.split():
                        cell.alignment = Alignment(horizontal='right')

            # adjust column width
            for column_cells in ws1.columns:
                length = max(len(cell.value or '') for cell in column_cells[1:])
                ws1.column_dimensions[column_cells[0].column].width = length * 1.5
            siaf_workbook.save(filename=out_filename)
            if verbose:
                print('Wrote Siaf to xlsx file {}'.format(out_filename))

        else:
            table = Table()
            for attribute_name in PRD_REQUIRED_ATTRIBUTES_ORDERED:
                data = [getattr(aperture_collection.apertures[aperture_name], attribute_name) for aperture_name in aperture_names]
                table.add_column(Column(data=data, name=attribute_name))
            table.write(out_filename, format=file_format)
            if verbose:
                print('Wrote Siaf to {} file {}'.format(file_format, out_filename))

        filenames.append(out_filename)

    return filenames
